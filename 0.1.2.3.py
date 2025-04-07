import urllib.request
from io import BytesIO
import gzip
from bs4 import BeautifulSoup
import re
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from datetime import datetime
import os
import time
import pandas as pd
from ebooklib import epub
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from tqdm import tqdm
import hashlib

# 正则表达式对象
findthreadid = re.compile(r'<li><a href="/Admin/Content/sagePost/id/(.*?).html">SAGE</a></li>')
findthreaduid = re.compile(r'<span class="h-threads-info-uid">ID:(.*?)</span>')
findcreatedat = re.compile(r'<span class="h-threads-info-createdat">(.*?)</span>')
findcontent = re.compile(r'<div class="h-threads-content">\n(.*?)</div>', re.S)
findmaintitle = re.compile(r'<span class="h-threads-info-title">(.*?)</span>')  # 匹配span标签
findmainemail = re.compile(r'<span class="h-threads-info-email">(.*?)</span>')
find_original_images = re.compile(r'https://image.nmb.best/image/.*?.(?:jpg|png|gif)')

def main():
    root = tk.Tk()
    root.withdraw()
    
    # 获取用户Cookie
    cookie = get_user_cookie()
    if not cookie:
        print("未提供Cookie，程序终止。")
        return
    
    # 选择包含串号的TXT文件
    file_path = filedialog.askopenfilename(title="选择包含串号的TXT文件", filetypes=[("Text files", "*.txt")])
    if not file_path:
        print("未选择文件，程序终止。")
        return
    
    # 读取串号列表
    with open(file_path, 'r', encoding='utf-8') as f:
        thread_ids = [line.strip() for line in f.readlines() if line.strip()]
    
    # 遍历处理每个串号
    for thread_id in thread_ids:
        print(f"\n开始处理串号：{thread_id}")
        try:
            process_thread(thread_id, cookie)
        except Exception as e:
            print(f"处理串号 {thread_id} 时发生错误：{str(e)}")

def get_user_cookie():
    """获取用户输入的Cookie，优先使用缓存"""
    CACHE_DIR = os.path.join(os.getenv('APPDATA'), 'CookieCache')
    COOKIE_PATH = os.path.join(CACHE_DIR, 'cookie.txt')
    os.makedirs(CACHE_DIR, exist_ok=True)

    old_cookie = None
    if os.path.exists(COOKIE_PATH):
        with open(COOKIE_PATH, 'r', encoding='utf-8') as f:
            old_cookie = f.read().strip()

    use_new = True
    if old_cookie:
        use_new = messagebox.askyesno("Cookie设置", "是否使用新Cookie？")

    if use_new:
        cookie = simpledialog.askstring("输入Cookie", "请输入Cookie值：")
        if cookie is None:
            messagebox.showinfo("提示", "未输入Cookie，程序终止。")
            return None
        while not cookie.strip():
            messagebox.showwarning("输入错误", "Cookie不能为空，请重新输入。")
            cookie = simpledialog.askstring("输入Cookie", "请输入Cookie值：")
            if cookie is None:
                messagebox.showinfo("提示", "未输入Cookie，程序终止。")
                return None
        with open(COOKIE_PATH, 'w', encoding='utf-8') as f:
            f.write(cookie.strip())
    else:
        cookie = old_cookie

    return cookie

def process_thread(thread_id, cookie):
    baseurl = f'https://www.nmbxd1.com/t/{thread_id}?page='
    datalist, author_id, main_title, main_email = get_data(baseurl, cookie)
    save_date = datetime.now().strftime('%Y%m%d')
    
    # 创建输出目录
    output_dir = f"{thread_id}_{save_date}"
    os.makedirs(output_dir, exist_ok=True)
    
    # 保存Excel文件（带图片）
    xlsx_path = os.path.join(output_dir, f"{main_title}_{save_date}.xlsx")
    save_data_to_xlsx(datalist, xlsx_path, main_title, main_email)
    
    # 生成EPUB文档
    epub_path = os.path.join(output_dir, f"{main_title}_{save_date}.epub")
    generate_epub(datalist, epub_path, thread_id, author_id, save_date, cookie)  # 添加cookie参数
    
    # 下载图片并插入Excel
    download_and_insert_images(datalist, output_dir, cookie, xlsx_path)

def ask_url(url, cookie):
    head = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.1",
        "Cookie": cookie
    }
    req = urllib.request.Request(url, headers=head)
    try:
        with urllib.request.urlopen(req) as response:
            h = response.read()
            buff = BytesIO(h)
            f = gzip.GzipFile(fileobj=buff)
            return f.read().decode('utf-8')
    except Exception as e:
        print(f"请求URL时发生错误：{str(e)}")
        return ""

def parse_page(html):
    soup = BeautifulSoup(html, "html.parser")
    main_item = soup.find('div', class_="h-threads-item-main")
    items = soup.find_all('div', class_="h-threads-item-reply-main")
    theme = soup.find('h2', class_="h-title")
    return main_item, items, theme

def parse_item(item):
    data = []
    item_str = str(item)
    
    # 提取基础信息
    threadid = re.findall(findthreadid, item_str)
    data.append(threadid[0] if threadid else "")
    
    threaduid = re.findall(findthreaduid, item_str)
    data.append(threaduid[0] if threaduid else "")
    
    createdat = re.findall(findcreatedat, item_str)
    data.append(createdat[0] if createdat else "")
    
    # 处理内容
    content = re.findall(findcontent, item_str)
    if content:
        content = content[0]
        content = re.sub(r'<.*?>', '', content)
        content = re.sub(r'&gt;', "＞", content)
        content = re.sub(r'&lt;', "＜", content)
        content = re.sub(r'<font color="#789922">', "", content)
        content = re.sub(r'</font>', "", content)
        data.append(content.strip())
    else:
        data.append("")
    
    # 修正：合并图片URL处理，仅保留去重后的部分
    # 提取并去重图片URL
    img_urls = []
    seen_urls = set()
    for url in re.findall(find_original_images, item_str):
        if url not in seen_urls:
            seen_urls.add(url)
            img_urls.append(url)
    data.append('; '.join(img_urls))  # 仅添加一次图片URL列
    
    return data

def get_data(baseurl, cookie):
    datalist = []
    page = 1
    author_id = main_title = main_email = ""
    
    while True:
        url = baseurl + str(page)
        html = ask_url(url, cookie)
        if not html:
            break
            
        main_item, items, theme = parse_page(html)
        
        if page == 1 and main_item:
            data = parse_item(main_item)
            datalist.append(data)
            author_id = data[1]
            
            # 从主项中提取标题和副标题（关键修改部分）
            main_title_match = re.findall(findmaintitle, str(main_item))
            main_title = main_title_match[0] if main_title_match else ""
            main_email_match = re.findall(findmainemail, str(main_item))
            main_email = main_email_match[0] if main_email_match else ""
        
        if page > 1 and len(items) <= 1:
            break
            
        for item in items:
            data = parse_item(item)
            if data[0] != "9999999":
                datalist.append(data)
                
        page += 1
        time.sleep(0)  # 控制爬取速度
        
    return datalist, author_id, main_title, main_email

def save_data_to_xlsx(datalist, save_path, main_title, main_email):
    # 仅保留前四列数据：串号、饼干、时间、内容
    df = pd.DataFrame([d[:4] for d in datalist], columns=["串号", "饼干", "时间", "内容"])
    
    # 使用上下文管理器确保正确保存
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='主数据')
        
        # 获取工作簿和工作表对象
        workbook = writer.book
        worksheet = writer.sheets['主数据']
        
        # 添加标题信息到F列
        worksheet.cell(row=1, column=5, value=main_title)  # F1单元格存储标题
        worksheet.cell(row=2, column=5, value=main_email)   # F2单元格存储副标题
        
        # 设置列宽
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 25
        worksheet.column_dimensions['D'].width = 60
        worksheet.column_dimensions['E'].width = 40  # 图片列宽度

    print(f"Excel文件已保存至：{save_path}")

def generate_epub(datalist, epub_path, thread_id, author_id, save_date, cookie):
    """生成包含图片的EPUB文件"""
    # ================== 初始化验证 ==================
    # 清理非法字符（关键修复）
    sanitized_title = re.sub(r'[\\/*?:"<>|]', "", str(thread_id))
    epub_path = os.path.join(
        os.path.dirname(epub_path),
        f"{sanitized_title}_{save_date}.epub"
    )
    print(f"EPUB最终保存路径：{epub_path}")

    # ================== 书籍初始化 ==================
    book = epub.EpubBook()
    book.set_identifier(f'xian_island_{thread_id}_{save_date}')
    book.set_title(f'X岛串号 {thread_id}')
    book.add_author('X岛匿名用户')
    book.set_language('zh-CN')

    # ================== 图片处理系统 ==================
    IMG_DIR = 'images'
    image_cache = {}

    def download_image(url):
        """带错误处理的图片下载"""
        if not url or url in image_cache:
            return image_cache.get(url)
        
        try:
            # 强化请求头（修复403错误）
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
                "Referer": "https://www.nmbxd1.com/",
                "Cookie": cookie
            }
            req = urllib.request.Request(url, headers=headers)
            
            with urllib.request.urlopen(req, timeout=15) as response:
                # 智能识别文件类型（修复扩展名错误）
                content_type = response.headers.get('Content-Type', 'image/jpeg')
                ext_map = {
                    'image/jpeg': 'jpg',
                    'image/png': 'png',
                    'image/gif': 'gif'
                }
                ext = ext_map.get(content_type.split(';')[0], 'jpg')
                
                # 生成哈希文件名（避免重复）
                file_hash = hashlib.md5(url.encode()).hexdigest()
                epub_img_path = f"{IMG_DIR}/{file_hash}.{ext}"
                
                # 创建图片对象
                img_item = epub.EpubImage(
                    uid=f"img_{len(image_cache)}",
                    file_name=epub_img_path,
                    media_type=content_type,
                    content=response.read()
                )
                book.add_item(img_item)
                image_cache[url] = epub_img_path
                return epub_img_path
        except Exception as e:
            print(f"⚠️ 图片下载失败：{url} - {str(e)}")
            image_cache[url] = None
            return None

    # ================== 图片预处理 ==================
    all_images = []
    seen_urls = set()
    for data in datalist:
        if data[4]:
            for url in data[4].split('; '):
                clean_url = url.strip()
                if clean_url and clean_url not in seen_urls:
                    seen_urls.add(clean_url)
                    all_images.append(clean_url)

    # ================== 下载进度条 ==================
    with tqdm(total=len(all_images), desc="下载图片资源", unit="img") as pbar:
        for url in all_images:
            download_image(url)
            pbar.update(1)

    # ================== 内容生成 ==================
    # 封面章节
    c_cover = epub.EpubHtml(title='封面', file_name='cover.xhtml', lang='zh-CN')
    c_cover.content = f'''
    <!DOCTYPE html>
    <html xmlns="http://www.w3.org/1999/xhtml">
    <head><title>封面</title></head>
    <body>
        <h1>{sanitized_title}</h1>
        <p>串号：{thread_id}</p>
        <p>作者ID：{author_id}</p>
        <p>存档时间：{save_date}</p>
    </body>
    </html>
    '''
    book.add_item(c_cover)

    # 主内容章节
    c_content = epub.EpubHtml(title='内容', file_name='content.xhtml', lang='zh-CN')
    posts_html = []
    
    for data in datalist:
        # 处理文本内容
        content = data[3].replace('\n', '<br/>')
        
        # 生成图片标签
        img_tags = []
        if data[4]:
            for url in data[4].split('; '):
                clean_url = url.strip()
                if not clean_url:
                    continue
                img_path = image_cache.get(clean_url)
                if img_path:
                    img_tags.append(f'<img src="{img_path}" class="content-image"/>')

        # 构建HTML结构
        post_html = f'''
        <article class="post">
            <header class="post-header">
                <span class="post-time">{data[2]}</span>
                <span class="post-id">ID:{data[1]}</span>
                {'''<span class="po-flag">(PO主)</span>''' if data[1] == author_id else ''}
            </header>
            <div class="post-content">
                <p>{content}</p>
                {'''<div class="images">''' + '\n'.join(img_tags) + '''</div>''' if img_tags else ''}
            </div>
            <hr/>
        </article>
        '''
        posts_html.append(post_html)

    c_content.content = f'''
    <!DOCTYPE html>
    <html xmlns="http://www.w3.org/1999/xhtml">
    <head>
        <title>{sanitized_title}</title>
        <link rel="stylesheet" type="text/css" href="style.css"/>
    </head>
    <body>
        <h2>{sanitized_title}</h2>
        {'\n'.join(posts_html)}
    </body>
    </html>
    '''
    book.add_item(c_content)

    # ================== 样式表 ==================
    style = '''
    body { font-family: "Microsoft YaHei", sans-serif; line-height: 1.6; }
    .post { margin-bottom: 2em; }
    .post-header { color: #666; border-left: 3px solid #789922; padding-left: 0.5em; }
    .post-time { font-weight: bold; }
    .post-id { color: #789922; margin-left: 1em; }
    .po-flag { color: #ff4444; }
    .content-image { max-width: 95%; height: auto; display: block; margin: 1em auto; }
    .images { text-align: center; }
    hr { border: 0; border-top: 1px dashed #ddd; margin: 2em 0; }
    '''
    css_item = epub.EpubItem(uid="style", file_name="style.css", media_type="text/css", content=style)
    book.add_item(css_item)

    # ================== 最终生成 ==================
    # 设置书籍结构
    book.toc = (epub.Link('cover.xhtml', '封面', 'cover'),
                epub.Link('content.xhtml', '内容', 'content'))
    book.spine = ['nav', c_cover, c_content]
    
    # 确保输出目录存在
    os.makedirs(os.path.dirname(epub_path), exist_ok=True)
    
    try:
        epub.write_epub(epub_path, book, {})
        print(f"✅ EPUB文件已成功生成：{epub_path}")
    except Exception as e:
        print(f"❌ EPUB保存失败：{str(e)}")
        if "Invalid characters" in str(e):
            print("错误原因：路径包含非法字符")
        raise

def download_and_insert_images(datalist, output_dir, cookie, xlsx_path):
    # 创建图片目录
    img_dir = os.path.join(output_dir, "images")
    os.makedirs(img_dir, exist_ok=True)
    
    # 全局记录已下载图片 {url: filename}
    downloaded = {}
    
    # 下载图片（带进度条）
    with tqdm(total=sum(len(re.findall(find_original_images, data[4])) for data in datalist),
              desc="下载图片") as pbar:
        for data in datalist:
            img_urls = re.findall(find_original_images, data[4])
            for url in img_urls:
                # 跳过已下载的
                if url in downloaded:
                    pbar.update(1)
                    continue
                
                try:
                    # 从URL提取纯净文件名（去除参数）
                    filename = os.path.basename(url.split('?')[0])
                    save_path = os.path.join(img_dir, filename)
                    
                    # 下载文件
                    req = urllib.request.Request(url, headers={"Cookie": cookie})
                    with urllib.request.urlopen(req) as response:
                        with open(save_path, 'wb') as f:
                            f.write(response.read())
                    
                    downloaded[url] = save_path
                except Exception as e:
                    print(f"下载失败：{url}，错误：{e}")
                    downloaded[url] = None  # 记录失败状态
                finally:
                    pbar.update(1)

    # 插入图片到Excel
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    
    # 添加图片列（如果不存在）
    if ws.cell(row=1, column=5).value != "图片":
        ws.insert_cols(5)
        ws.cell(row=1, column=5, value="图片")
    
    # 插入图片路径和缩略图
    for row_idx, data in enumerate(datalist, start=2):
        img_urls = re.findall(find_original_images, data[4])
        valid_paths = []
        
        # 获取有效图片路径
        for url in img_urls:
            path = downloaded.get(url)
            if path and os.path.exists(path):
                valid_paths.append(path)
        
        # 写入单元格
        if valid_paths:
            cell = ws.cell(row=row_idx, column=5)
            cell.value = "\n".join(valid_paths)
            
            # 插入第一张有效图片
            try:
                img = ExcelImage(valid_paths[0])
                img.width = 100
                img.height = 100
                ws.add_image(img, f"E{row_idx}")
            except Exception as e:
                print(f"插入图片失败：{valid_paths[0]} - {e}")

    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 40
    
    wb.save(xlsx_path)
    print(f"图片已插入Excel文件：{xlsx_path}")

if __name__ == "__main__":
    main()
    print("所有串号处理完毕！")
