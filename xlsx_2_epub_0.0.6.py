import os
import re
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import urllib.request
from ebooklib import epub
import hashlib
from tqdm import tqdm

def main():
    root = tk.Tk()
    root.withdraw()
    
    # 选择XLSX文件
    xlsx_path = filedialog.askopenfilename(
        title="选择爬虫生成的XLSX文件",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not xlsx_path:
        print("未选择文件，程序终止。")
        return

    try:
        # 新增：明确定义工作表对象
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        
        # 增强版数据获取（带空值检查）
        main_title = ws['F1'].value if ws['F1'].value else "A2"
        main_email = ws['F2'].value if ws['F2'].value else ""
        author_id = ws['B2'].value  # 获取第一个发言的饼干ID

        # 如果作者ID为空时的处理
        if not author_id:
            print("错误：未能识别PO主饼干，请确保文件包含有效发言数据")
            return

        # 读取所有PO主发言
        po_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == author_id:  # B列匹配作者ID
                content = row[3]     # D列内容
                images = row[4] if len(row) > 4 else ""  # E列图片
                po_data.append({
                    "content": content,
                    "images": images.split('; ') if images else []
                })

        # 生成EPUB
        output_dir = os.path.join(os.path.dirname(xlsx_path), "EPUB输出")
        os.makedirs(output_dir, exist_ok=True)
        
        sanitized_title = re.sub(r'[\\/*?:"<>|]', "", str(main_title))
        epub_path = os.path.join(output_dir, f"{sanitized_title}.epub")
        
        generate_po_epub(po_data, epub_path, main_title, author_id, main_email)

    except Exception as e:
        print(f"处理文件时发生严重错误: {str(e)}")

def generate_po_epub(data, epub_path, main_title, author_id, main_email):
    """生成符合新排版要求的EPUB"""
    # 安全处理输入参数
    safe_title = str(main_title or "未知串号").strip()
    safe_subtitle = str(main_email or "").strip()
    safe_author = str(author_id or "未知饼干").strip()

    # 书籍初始化
    book = epub.EpubBook()
    book.set_identifier(f'po_{hashlib.md5(safe_title.encode()).hexdigest()}')
    book.set_title(safe_title)
    book.add_author(safe_author)
    book.set_language('zh-CN')

    # ================== 图片处理系统 ==================
    IMG_DIR = 'images'
    image_cache = {}
    
    def download_image(url):
        """增强型图片下载"""
        if not url or url in image_cache:
            return image_cache.get(url)
        
        try:
            # 生成安全文件名
            file_hash = hashlib.md5(url.encode()).hexdigest()
            ext = os.path.splitext(url)[1].split('?')[0][1:4]  # 处理带参数的URL
            epub_img_path = f"{IMG_DIR}/{file_hash}.{ext or 'jpg'}"
            
            # 下载并缓存
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Referer": "https://www.nmbxd1.com/"
            })
            with urllib.request.urlopen(req, timeout=15) as res:
                img_item = epub.EpubImage(
                    uid=f"img_{len(image_cache)}",
                    file_name=epub_img_path,
                    media_type=res.headers['Content-Type'],
                    content=res.read()
                )
                book.add_item(img_item)
                image_cache[url] = epub_img_path
                return epub_img_path
        except Exception as e:
            print(f"图片下载失败：{url} - {str(e)}")
            image_cache[url] = None
            return None

    # ================== 预处理所有图片 ==================
    all_images = []
    for item in data:
        all_images.extend(item["images"])
    
    with tqdm(total=len(all_images), desc="预处理图片资源") as pbar:
        seen = set()
        unique_images = [url for url in all_images if not (url in seen or seen.add(url))]
        for url in unique_images:
            download_image(url)
            pbar.update(1)

    # ================== 内容生成 ==================
    c_content = epub.EpubHtml(title='内容', file_name='content.xhtml')
    
    # 构建标题区块（更新版）
    title_block = [
        '<div class="title-section">',
        f'<h1 class="main-title">{safe_title}</h1>',
    ]
    
    if safe_subtitle:
        title_block.append(
            f'<h2 class="subtitle">{safe_subtitle}</h2>'
        )
    
    title_block.extend([
        f'<div class="author-info">作者：{safe_author}</div>',
        '</div>',
        '<hr class="title-separator"/>'
    ])

    # 构建正文内容
    content_blocks = []
    for item in data:
        # 处理文本内容（保留换行）
        clean_content = item["content"].replace('\n', '<br/>')
        
        # 生成图片标签
        img_tags = []
        for url in item["images"]:
            if img_path := image_cache.get(url):
                img_tags.append(f'<img src="{img_path}" style="max-width: 90%; height: auto; display: block; margin: 1em auto;"/>')
        
        # 构建内容块
        block = f'''
        <section class="po-post">
            <div class="content-text">{clean_content}</div>
            {'''<div class="images">''' + '\n'.join(img_tags) + '''</div>''' if img_tags else ''}
        </section>
        <hr/>
        '''
        content_blocks.append(block)

    c_content.content = f'''
    <!DOCTYPE html>
    <html>
    <head>
        <title>{safe_title}</title>
        <style>
            /* 更新后的CSS样式 */
            body {{ 
                font-family: "Microsoft YaHei", sans-serif; 
                line-height: 1.8;
                margin: 2em;
            }}
            .title-section {{
                text-align: center;
                margin: 3em 0;
            }}
            .main-title {{
                font-size: 2.2em;
                margin: 0.5em 0;
                color: #2c3e50;
            }}
            .subtitle {{
                font-size: 1.4em;
                color: #7f8c8d;
                margin: 1em 0;
                font-weight: normal;
            }}
            .author-info {{
                font-size: 1.1em;
                color: #95a5a6;
                margin-top: 1.5em;
            }}
            .title-separator {{
                border: 0;
                height: 2px;
                background: #bdc3c7;
                margin: 2em 0;
            }}
            /* 原有其他样式保持不变 */
        </style>
    </head>
    <body>
        {"".join(title_block)}
        {''.join(content_blocks)}
    </body>
    </html>
    '''
    book.add_item(c_content)

    # ================== 生成最终文件 ==================
    book.toc = (epub.Link('content.xhtml', '内容', 'content'),)
    book.spine = ['nav', c_content]
    
    try:
        epub.write_epub(epub_path, book, {})
        print(f"EPUB生成成功：{epub_path}")
    except Exception as e:
        print(f"生成失败：{str(e)}")
        if "Invalid characters" in str(e):
            print("提示：请检查标题是否包含特殊字符")

if __name__ == "__main__":
    main()