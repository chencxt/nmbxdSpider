import urllib.request
import gzip
from bs4 import BeautifulSoup
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
import os
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import threading
import queue

# 正则表达式对象
findthreadid = re.compile(r'<li><a href="/Admin/Content/sagePost/id/(.*?).html">SAGE</a></li>')
findthreaduid = re.compile(r'<span class="h-threads-info-uid">ID:(.*?)</span>')
findcreatedat = re.compile(r'<span class="h-threads-info-createdat">(.*?)</span>')
findcontent = re.compile(r'<div class="h-threads-content">\n(.*?)</div>', re.S)
findmaintitle = re.compile(r'<span class="h-threads-info-title">(.*?)</span>')
findmainemail = re.compile(r'<span class="h-threads-info-email">(.*?)</span>')
find_original_images = re.compile(r'https://image.nmb.best/image/.*?.(?:jpg|png|gif)')

class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("X岛爬虫 v0.2.0.1")
        self.geometry("800x600")
        self.queue = queue.Queue()
        self.create_widgets()
        self.load_cached_cookie()
        self.after(100, self.process_queue)

    def create_widgets(self):
        mainframe = ttk.Frame(self, padding=10)
        mainframe.pack(fill=tk.BOTH, expand=True)

        # Cookie输入
        ttk.Label(mainframe, text="Cookie:").grid(row=0, column=0, sticky=tk.W)
        self.cookie_entry = ttk.Entry(mainframe, width=50)
        self.cookie_entry.grid(row=0, column=1, sticky=tk.EW, padx=5)

        # 文件选择
        ttk.Label(mainframe, text="串号文件:").grid(row=1, column=0, sticky=tk.W)
        self.file_path = tk.StringVar()
        ttk.Entry(mainframe, textvariable=self.file_path, state='readonly').grid(row=1, column=1, sticky=tk.EW, padx=5)
        ttk.Button(mainframe, text="浏览", command=self.browse_file).grid(row=1, column=2, padx=5)

        # 进度条
        self.progress = ttk.Progressbar(mainframe, orient=tk.HORIZONTAL, mode='determinate')
        self.progress.grid(row=2, column=0, columnspan=3, sticky=tk.EW, pady=5)

        # 分页进度
        self.page_status = ttk.Label(mainframe, text="就绪")
        self.page_status.grid(row=3, column=0, columnspan=3, sticky=tk.W)

        # 日志区域
        log_frame = ttk.Frame(mainframe)
        log_frame.grid(row=4, column=0, columnspan=3, sticky=tk.NSEW)
        self.log_text = tk.Text(log_frame, wrap=tk.WORD, height=15)
        vsb = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=vsb.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # 操作按钮
        ttk.Button(mainframe, text="开始处理", command=self.start_processing).grid(row=5, column=1, pady=10)

        mainframe.columnconfigure(1, weight=1)
        mainframe.rowconfigure(4, weight=1)

    def browse_file(self):
        filepath = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if filepath:
            self.file_path.set(filepath)

    def load_cached_cookie(self):
        cache_dir = os.path.join(os.getenv('APPDATA'), 'CookieCache')
        cookie_path = os.path.join(cache_dir, 'cookie.txt')
        if os.path.exists(cookie_path):
            with open(cookie_path, 'r') as f:
                self.cookie_entry.insert(0, f.read().strip())

    def start_processing(self):
        cookie = self.cookie_entry.get().strip()
        filepath = self.file_path.get()

        if not cookie:
            messagebox.showerror("错误", "请输入Cookie！")
            return
        if not filepath:
            messagebox.showerror("错误", "请选择串号文件！")
            return

        try:
            with open(filepath, 'r') as f:
                thread_ids = [line.strip() for line in f if line.strip()]
        except Exception as e:
            messagebox.showerror("错误", f"文件读取失败: {str(e)}")
            return

        self.progress['maximum'] = len(thread_ids)
        self.progress['value'] = 0
        self.log_text.delete(1.0, tk.END)
        self.page_status.config(text="就绪")

        threading.Thread(
            target=self.process_threads,
            args=(thread_ids, cookie),
            daemon=True
        ).start()

    def process_threads(self, thread_ids, cookie):
        try:
            cache_dir = os.path.join(os.getenv('APPDATA'), 'CookieCache')
            os.makedirs(cache_dir, exist_ok=True)
            with open(os.path.join(cache_dir, 'cookie.txt'), 'w') as f:
                f.write(cookie)

            for idx, tid in enumerate(thread_ids, 1):
                self.queue.put(('progress', idx))
                self.queue.put(('log', f"=== 开始处理串号: {tid} ==="))
                try:
                    self.process_single_thread(tid, cookie)
                except Exception as e:
                    self.queue.put(('log', f"处理失败: {str(e)}"))
                finally:
                    self.queue.put(('page_status', "准备下一个串号..."))
            
            self.queue.put(('log', "所有处理已完成！"))
        except Exception as e:
            self.queue.put(('log', f"处理过程中发生错误: {str(e)}"))
        finally:
            self.queue.put(('page_status', "就绪"))

    def process_single_thread(self, thread_id, cookie):
        baseurl = f'https://www.nmbxd1.com/t/{thread_id}?page='
        self.queue.put(('page_status', f"初始化处理环境: {thread_id}"))
        
        datalist, author_id, main_title, main_email = self.get_data(thread_id, baseurl, cookie)
        save_date = datetime.now().strftime('%Y%m%d')
        output_dir = f"{thread_id}_{save_date}"
        os.makedirs(output_dir, exist_ok=True)

        xlsx_path = os.path.join(output_dir, f"{main_title}_{save_date}.xlsx")
        self.save_data_to_xlsx(datalist, xlsx_path, main_title, main_email)
        self.download_and_insert_images(datalist, output_dir, cookie, xlsx_path)

    def get_data(self, thread_id, baseurl, cookie):
        datalist = []
        page = 1
        author_id = main_title = main_email = ""
        total_pages = 0
        total_replies = 0

        while True:
            url = baseurl + str(page)
            self.queue.put(('page_status', f"串号 {thread_id} - 请求第{page}页"))
            html = self.ask_url(url, cookie)
            
            main_item, items, theme = self.parse_page(html)
            reply_count = 0

            if page == 1 and main_item:
                data = self.parse_item(main_item)
                datalist.append(data)
                author_id = data[1]
                main_title_match = re.findall(findmaintitle, str(main_item))
                main_title = main_title_match[0] if main_title_match else ""
                main_email_match = re.findall(findmainemail, str(main_item))
                main_email = main_email_match[0] if main_email_match else ""

            valid_items = [item for item in items if self.parse_item(item)[0] != "9999999"]
            for item in valid_items:
                datalist.append(self.parse_item(item))
                reply_count += 1
                total_replies += 1

            self.queue.put(('log', f"串号 {thread_id} - 第{page}页"))
            self.queue.put(('page_status', f"串号 {thread_id} - 第{page}页 ({reply_count}条回复)"))

            if page > 1 and len(items) <= 1:
                self.queue.put(('log', f"串号 {thread_id} - 检测到最后一页，终止采集"))
                break
            page += 1
            total_pages += 1

        self.queue.put(('log', f"串号 {thread_id} 采集完成！总页数：{total_pages}，总回复：{total_replies}"))
        return datalist, author_id, main_title, main_email

    def ask_url(self, url, cookie):
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.1",
            "Cookie": cookie
        }
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=10) as response:
                return gzip.decompress(response.read()).decode('utf-8')
        except Exception as e:
            self.queue.put(('log', f"请求失败: {str(e)}"))
            return ""

    def parse_page(self, html):
        soup = BeautifulSoup(html, "html.parser")
        main_item = soup.find('div', class_="h-threads-item-main")
        items = soup.find_all('div', class_="h-threads-item-reply-main")
        theme = soup.find('h2', class_="h-title")
        return main_item, items, theme

    def parse_item(self, item):
        item_str = str(item)
        threadid = re.findall(findthreadid, item_str)
        threaduid = re.findall(findthreaduid, item_str)
        createdat = re.findall(findcreatedat, item_str)
        
        content = re.findall(findcontent, item_str)
        if content:
            content = re.sub(r'<.*?>', '', content[0])
            content = re.sub(r'&gt;', "＞", content)
            content = re.sub(r'&lt;', "＜", content)
            content = content.strip()
        else:
            content = ""
        
        img_urls = list(set(re.findall(find_original_images, item_str)))
        return [
            threadid[0] if threadid else "",
            threaduid[0] if threaduid else "",
            createdat[0] if createdat else "",
            content,
            '; '.join(img_urls)
        ]

    def save_data_to_xlsx(self, datalist, save_path, main_title, main_email):
        df = pd.DataFrame([d[:4] for d in datalist], columns=["串号", "饼干", "时间", "内容"])
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='主数据')
            workbook = writer.book
            worksheet = writer.sheets['主数据']
            worksheet.cell(row=1, column=5, value=main_title)
            worksheet.cell(row=2, column=5, value=main_email)
            for col_idx, width in enumerate([15, 20, 25, 60, 40], 1):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    def download_and_insert_images(self, datalist, output_dir, cookie, xlsx_path):
        img_dir = os.path.join(output_dir, "images")
        os.makedirs(img_dir, exist_ok=True)
        downloaded = {}
        total_images = sum(len(re.findall(find_original_images, data[4])) for data in datalist)
        processed = 0

        for data in datalist:
            for url in re.findall(find_original_images, data[4]):
                if url not in downloaded:
                    try:
                        self.queue.put(('page_status', f"下载图片 {processed+1}/{total_images}"))
                        filename = os.path.basename(url.split('?')[0])
                        save_path = os.path.join(img_dir, filename)
                        req = urllib.request.Request(url, headers={"Cookie": cookie})
                        with urllib.request.urlopen(req, timeout=10) as response:
                            with open(save_path, 'wb') as f:
                                f.write(response.read())
                        downloaded[url] = save_path
                    except Exception as e:
                        downloaded[url] = None
                        self.queue.put(('log', f"图片下载失败: {url} - {str(e)}"))
                    finally:
                        processed += 1

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        if ws.cell(row=1, column=5).value != "图片":
            ws.insert_cols(5)
            ws.cell(row=1, column=5, value="图片")

        for row_idx, data in enumerate(datalist, 2):
            img_paths = [downloaded[url] for url in re.findall(find_original_images, data[4]) if downloaded.get(url)]
            if img_paths:
                ws.cell(row=row_idx, column=5).value = "\n".join(img_paths)
                try:
                    img = ExcelImage(img_paths[0])
                    img.width, img.height = 100, 100
                    ws.add_image(img, f"E{row_idx}")
                except:
                    pass
        wb.save(xlsx_path)

    def process_queue(self):
        try:
            while True:
                msg = self.queue.get_nowait()
                if isinstance(msg, tuple):
                    if msg[0] == 'progress':
                        self.progress['value'] = msg[1]
                        self.title(f"X岛爬虫 v0.2.0.1 - 进度 {msg[1]}/{self.progress['maximum']}")
                    elif msg[0] == 'page_status':
                        self.page_status.config(text=msg[1])
                    elif msg[0] == 'log':
                        self.log_text.insert(tk.END, f"{msg[1]}\n")
                        self.log_text.see(tk.END)
        except queue.Empty:
            pass
        finally:
            self.after(100, self.process_queue)

if __name__ == "__main__":
    app = Application()
    app.mainloop()